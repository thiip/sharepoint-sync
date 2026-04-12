import os, json, io, time, logging, requests, unicodedata, re
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from msal import ConfidentialClientApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger(__name__)

# === CONFIG ===
TENANT_ID = os.environ["AZURE_TENANT_ID"]
CLIENT_ID = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
SHAREPOINT_SITE = os.environ.get("SHAREPOINT_SITE", "projectumm-my.sharepoint.com")
SHAREPOINT_PATH = os.environ.get("SHAREPOINT_PATH", "/personal/lucas_projectum_com_br1")
FILE_PATH = os.environ.get("FILE_PATH", "/Documents/Contabilidade reforma galpão.xlsx")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://supa.projectum.com.br")
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
SYNC_INTERVAL = int(os.environ.get("SYNC_INTERVAL_MINUTES", "30"))
SUPABASE_SCHEMA = os.environ.get("SUPABASE_SCHEMA", "galpao")

HEADERS_SB = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    # PostgREST schema routing — read e write precisam do mesmo schema
    "Accept-Profile": SUPABASE_SCHEMA,
    "Content-Profile": SUPABASE_SCHEMA,
}

# === NORMALIZATION ===
def _strip_accents(s):
    """Remove accents: Decoração → Decoracao"""
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))

def _norm(s):
    """Robust normalization for dedup matching:
    - strip + uppercase
    - remove accents
    - collapse multiple spaces
    - normalize punctuation (remove hyphens, dots, extra chars)
    """
    s = str(s or "").strip().upper()
    s = _strip_accents(s)
    s = re.sub(r'[-./]', ' ', s)       # replace punctuation with space
    s = re.sub(r'\s+', ' ', s).strip()  # collapse multiple spaces
    return s

def _norm_valor(v):
    """Normalize numeric value to a deterministic string for comparison.
    Uses Decimal to avoid float representation issues.
    """
    try:
        d = Decimal(str(v)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return str(d)
    except Exception:
        return str(round(float(v), 2))

# Supplier alias map: map known variant names to canonical form
SUPPLIER_ALIASES = {
    "HAVA PRODUTOS": "HAVA PRODUTOS QUIMICOS LTDA",
    "HAVA PRODUTOS QUIMICOS": "HAVA PRODUTOS QUIMICOS LTDA",
    "MG MUNK": "MG MUNCK",
    "CONTAGEMQUIPE": "CONTAGEMQUIP",
    "CONTAGEMQUIP": "CONTAGEMQUIP",
    "TELEPAR": "TELEPAR PARAFUSOS",
    "LT DECORACAO": "LT DECORACOES",
    "LT DECORACOES": "LT DECORACOES",
    "LT DECORACOES": "LT DECORACOES",
}

# Only two companies exist as payers
def _canonical_payer(p):
    """Normalize payer name: only 'LT Decoraçoes' or 'Just Smile'."""
    n = _strip_accents(str(p or "")).strip().lower()
    if "just" in n or "smile" in n:
        return "Just Smile"
    return "LT Decoraçoes"

def _canonical_supplier(name):
    """Resolve supplier name to canonical form."""
    normed = _norm(name)
    return SUPPLIER_ALIASES.get(normed, normed)

# === MICROSOFT GRAPH AUTH ===
def get_graph_token():
    app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise Exception(f"Auth failed: {result.get('error_description', result)}")
    return result["access_token"]

# === DOWNLOAD EXCEL FROM SHAREPOINT ===
def download_excel(token):
    headers = {"Authorization": f"Bearer {token}"}
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE}:{SHAREPOINT_PATH}"
    resp = requests.get(site_url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"Site not found: {resp.status_code}")
    site_id = resp.json()["id"]
    log.info(f"Site ID: {site_id}")

    search_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='Contabilidade reforma')"
    resp = requests.get(search_url, headers=headers)
    items = resp.json().get("value", [])
    if not items:
        raise Exception("Excel file not found in SharePoint")

    item_id = items[0]["id"]
    log.info(f"File: {items[0]['name']} (ID: {item_id})")

    content_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{item_id}/content"
    resp = requests.get(content_url, headers=headers)
    if resp.status_code == 200:
        return resp.content, site_id, item_id
    else:
        raise Exception(f"Download failed: {resp.status_code}")

# === UPLOAD EXCEL TO SHAREPOINT ===
def upload_excel(token, site_id, item_id, content_bytes):
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{item_id}/content"
    resp = requests.put(url, headers=headers, data=content_bytes)
    if resp.status_code in (200, 201):
        log.info("Excel uploaded to SharePoint successfully")
        return True
    else:
        log.error(f"Upload failed: {resp.status_code} - {resp.text[:300]}")
        return False

# === READ ERP ITEMS FROM SUPABASE ===
def read_erp_items():
    """Buscar registros criados/modificados pelo ERP (source='erp')"""
    despesas = []
    outros = []
    deleted_despesas = []
    deleted_outros = []

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/despesas?source=eq.erp&select=*",
        headers=HEADERS_SB,
    )
    if resp.status_code == 200:
        despesas = resp.json()
    log.info(f"ERP despesas pendentes: {len(despesas)}")

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/outros?source=eq.erp&select=*",
        headers=HEADERS_SB,
    )
    if resp.status_code == 200:
        outros = resp.json()
    log.info(f"ERP outros pendentes: {len(outros)}")

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/despesas?source=eq.erp_deleted&select=*",
        headers=HEADERS_SB,
    )
    if resp.status_code == 200:
        deleted_despesas = resp.json()
    log.info(f"ERP despesas deletadas pendentes: {len(deleted_despesas)}")

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/outros?source=eq.erp_deleted&select=*",
        headers=HEADERS_SB,
    )
    if resp.status_code == 200:
        deleted_outros = resp.json()
    log.info(f"ERP outros deletados pendentes: {len(deleted_outros)}")

    return despesas, outros, deleted_despesas, deleted_outros

# === WRITE ERP ITEMS TO EXCEL ===
def write_to_excel(wb, erp_despesas, erp_outros, deleted_despesas, deleted_outros):
    """Escrever itens do ERP no Excel e remover itens deletados"""
    modified = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if "GALP" in sheet_name.upper() or "OBRA" in sheet_name.upper():
            # --- Remover linhas deletadas ---
            if deleted_despesas:
                rows_to_delete = []
                for row in range(4, ws.max_row + 1):
                    desc = str(ws.cell(row=row, column=2).value or "")
                    data_cell = ws.cell(row=row, column=4).value
                    valor = ws.cell(row=row, column=6).value

                    if isinstance(data_cell, datetime):
                        data_str = data_cell.strftime("%Y-%m-%d")
                    elif data_cell:
                        data_str = str(data_cell)[:10]
                    else:
                        continue

                    try:
                        valor_f = float(valor) if valor else 0
                    except (ValueError, TypeError):
                        continue

                    for dd in deleted_despesas:
                        if (_norm(dd.get("descricao", "")) == _norm(desc) and
                            dd.get("data", "") == data_str and
                            abs(float(dd.get("valor", 0)) - valor_f) < 0.01):
                            rows_to_delete.append(row)
                            break

                for row in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row)
                    modified = True
                    log.info(f"  Deleted row {row} from sheet '{sheet_name}'")

            # --- Adicionar novos itens do ERP ---
            if erp_despesas:
                last_row = ws.max_row
                for row in range(ws.max_row, 3, -1):
                    if ws.cell(row=row, column=6).value is not None:
                        last_row = row
                        break

                for item in erp_despesas:
                    exists = False
                    for row in range(4, last_row + 1):
                        desc = str(ws.cell(row=row, column=2).value or "")
                        data_cell = ws.cell(row=row, column=4).value
                        valor = ws.cell(row=row, column=6).value
                        if isinstance(data_cell, datetime):
                            ds = data_cell.strftime("%Y-%m-%d")
                        elif data_cell:
                            ds = str(data_cell)[:10]
                        else:
                            continue
                        try:
                            valor_f = float(valor) if valor else 0
                        except (ValueError, TypeError):
                            continue
                        if (_norm(item.get("descricao", "")) == _norm(desc) and
                            item.get("data", "") == ds and
                            abs(float(item.get("valor", 0)) - valor_f) < 0.01):
                            exists = True
                            break

                    if not exists:
                        last_row += 1
                        ws.cell(row=last_row, column=2, value=item.get("descricao", ""))
                        ws.cell(row=last_row, column=3, value=item.get("obs", ""))

                        try:
                            dt = datetime.strptime(item["data"], "%Y-%m-%d")
                            cell_d = ws.cell(row=last_row, column=4, value=dt)
                            cell_d.number_format = 'DD/MM/YYYY'
                        except (ValueError, KeyError):
                            ws.cell(row=last_row, column=4, value=item.get("data", ""))

                        ws.cell(row=last_row, column=5, value=item.get("pago", ""))

                        cell_v = ws.cell(row=last_row, column=6, value=float(item.get("valor", 0)))
                        cell_v.number_format = '#,##0.00'

                        modified = True
                        log.info(f"  Added to Excel: {item.get('descricao')} - R${item.get('valor')}")

        elif "OUTROS" in sheet_name.upper():
            sections = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=2, column=col).value
                if val and str(val).strip():
                    cat_name = str(val).strip()
                    data_col = None
                    valor_col = None
                    for c in range(col, min(col + 4, ws.max_column + 1)):
                        h = str(ws.cell(row=3, column=c).value or "").strip().upper()
                        if h == "DATA":
                            data_col = c
                        elif h == "VALOR":
                            valor_col = c
                    if data_col and valor_col:
                        sections.append({"cat": cat_name, "data_col": data_col, "valor_col": valor_col})

            if deleted_outros:
                for section in sections:
                    rows_to_clear = []
                    for row in range(4, ws.max_row + 1):
                        data_cell = ws.cell(row=row, column=section["data_col"]).value
                        valor = ws.cell(row=row, column=section["valor_col"]).value
                        if data_cell is None or valor is None:
                            continue
                        if isinstance(data_cell, datetime):
                            ds = data_cell.strftime("%Y-%m-%d")
                        else:
                            ds = str(data_cell)[:10]
                        try:
                            valor_f = float(valor)
                        except (ValueError, TypeError):
                            continue
                        for do in deleted_outros:
                            if (do.get("cat", "") == section["cat"] and
                                do.get("data", "") == ds and
                                abs(float(do.get("valor", 0)) - valor_f) < 0.01):
                                rows_to_clear.append(row)
                                break

                    for row in rows_to_clear:
                        ws.cell(row=row, column=section["data_col"]).value = None
                        ws.cell(row=row, column=section["valor_col"]).value = None
                        modified = True
                        log.info(f"  Cleared OUTROS row {row} cat={section['cat']}")

            if erp_outros:
                for item in erp_outros:
                    target_section = None
                    for s in sections:
                        if s["cat"] == item.get("cat", ""):
                            target_section = s
                            break

                    if not target_section:
                        log.warning(f"  Categoria '{item.get('cat')}' não encontrada no Excel OUTROS")
                        continue

                    exists = False
                    last_row = 3
                    for row in range(4, ws.max_row + 1):
                        data_cell = ws.cell(row=row, column=target_section["data_col"]).value
                        valor = ws.cell(row=row, column=target_section["valor_col"]).value
                        if data_cell is not None and valor is not None:
                            last_row = row
                            if isinstance(data_cell, datetime):
                                ds = data_cell.strftime("%Y-%m-%d")
                            else:
                                ds = str(data_cell)[:10]
                            try:
                                v = float(valor)
                            except (ValueError, TypeError):
                                continue
                            if (item.get("data", "") == ds and
                                abs(float(item.get("valor", 0)) - v) < 0.01):
                                exists = True

                    if not exists:
                        new_row = last_row + 1
                        try:
                            dt = datetime.strptime(item["data"], "%Y-%m-%d")
                            cell_d = ws.cell(row=new_row, column=target_section["data_col"], value=dt)
                            cell_d.number_format = 'DD/MM/YYYY'
                        except (ValueError, KeyError):
                            ws.cell(row=new_row, column=target_section["data_col"], value=item.get("data", ""))

                        cell_v = ws.cell(row=new_row, column=target_section["valor_col"], value=float(item.get("valor", 0)))
                        cell_v.number_format = '#,##0.00'

                        modified = True
                        log.info(f"  Added to Excel OUTROS: {item.get('cat')} - R${item.get('valor')}")

    return modified

# === MARK ERP ITEMS AS SYNCED ===
def mark_erp_items_synced():
    """Marcar itens do ERP como sincronizados (source='synced')"""
    resp = requests.patch(
        f"{SUPABASE_URL}/rest/v1/despesas?source=eq.erp",
        headers=HEADERS_SB,
        json={"source": "synced"},
    )
    if resp.status_code < 300:
        log.info("Despesas ERP marcadas como synced")
    else:
        log.error(f"Erro ao marcar despesas: {resp.status_code}")

    resp = requests.patch(
        f"{SUPABASE_URL}/rest/v1/outros?source=eq.erp",
        headers=HEADERS_SB,
        json={"source": "synced"},
    )
    if resp.status_code < 300:
        log.info("Outros ERP marcados como synced")
    else:
        log.error(f"Erro ao marcar outros: {resp.status_code}")

    resp = requests.delete(
        f"{SUPABASE_URL}/rest/v1/despesas?source=eq.erp_deleted",
        headers=HEADERS_SB,
    )
    log.info(f"Despesas erp_deleted removidas: {resp.status_code}")

    resp = requests.delete(
        f"{SUPABASE_URL}/rest/v1/outros?source=eq.erp_deleted",
        headers=HEADERS_SB,
    )
    log.info(f"Outros erp_deleted removidos: {resp.status_code}")

# === PARSE EXCEL ===
def parse_excel(content):
    wb = load_workbook(io.BytesIO(content), data_only=True)
    despesas = []
    outros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.info(f"Sheet: '{sheet_name}' ({ws.max_row} rows x {ws.max_column} cols)")

        if "GALP" in sheet_name.upper() or "OBRA" in sheet_name.upper():
            for row in range(4, ws.max_row + 1):
                valor = ws.cell(row=row, column=6).value
                if valor is None or valor == "" or valor == 0:
                    continue
                try:
                    valor = float(valor)
                except (ValueError, TypeError):
                    continue

                data_cell = ws.cell(row=row, column=4).value
                if isinstance(data_cell, datetime):
                    data_str = data_cell.strftime("%Y-%m-%d")
                elif data_cell:
                    data_str = str(data_cell)[:10]
                else:
                    continue

                despesas.append({
                    "descricao": str(ws.cell(row=row, column=2).value or "").strip().upper(),
                    "obs": str(ws.cell(row=row, column=3).value or "").strip(),
                    "data": data_str,
                    "pago": _canonical_payer(ws.cell(row=row, column=5).value),
                    "valor": valor,
                })

        elif "OUTROS" in sheet_name.upper():
            sections = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=2, column=col).value
                if val and str(val).strip():
                    cat_name = str(val).strip()
                    data_col = None
                    valor_col = None
                    for c in range(col, min(col + 4, ws.max_column + 1)):
                        h = str(ws.cell(row=3, column=c).value or "").strip().upper()
                        if h == "DATA":
                            data_col = c
                        elif h == "VALOR":
                            valor_col = c
                    if data_col and valor_col:
                        sections.append({"cat": cat_name, "data_col": data_col, "valor_col": valor_col})

            log.info(f"  OUTROS sections: {[s['cat'] for s in sections]}")

            for section in sections:
                for row in range(4, ws.max_row + 1):
                    valor = ws.cell(row=row, column=section["valor_col"]).value
                    if valor is None or valor == "" or valor == 0:
                        continue
                    try:
                        valor = float(valor)
                    except (ValueError, TypeError):
                        continue

                    data_cell = ws.cell(row=row, column=section["data_col"]).value
                    if isinstance(data_cell, datetime):
                        data_str = data_cell.strftime("%Y-%m-%d")
                    elif data_cell:
                        data_str = str(data_cell)[:10]
                        if not data_str[:4].isdigit():
                            continue
                    else:
                        continue

                    outros.append({
                        "cat": section["cat"],
                        "data": data_str,
                        "valor": valor,
                    })

    log.info(f"Parsed: {len(despesas)} despesas, {len(outros)} outros")
    return despesas, outros

# === SYNC TO SUPABASE (COUNTER-BASED DEDUP) ===
def sync_to_supabase(despesas, outros):
    """Idempotent sync using COUNTER-BASED dedup.

    The Excel can have duplicate rows that are legitimate (e.g. two identical
    payments on the same day). We count how many times each key exists in the
    DB and in the incoming Excel batch. We only insert the DIFFERENCE.

    This guarantees: DB record count per key == Excel row count per key.
    """
    from collections import Counter
    stats = {"inserted": 0, "skipped": 0, "errors": 0}

    # ── DESPESAS ──────────────────────────────────────────────────────────────
    if despesas:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/despesas?select=id,descricao,data,valor,pago,source"
            f"&source=neq.erp_deleted",
            headers=HEADERS_SB,
        )
        existing = resp.json() if resp.status_code == 200 else []

        # Count existing records per key in DB
        db_counts = Counter()
        for e in existing:
            key = (_norm(e["descricao"]), e["data"], _norm_valor(e["valor"]),
                   _norm(e["pago"]))
            db_counts[key] += 1

        # Count incoming records per key from Excel
        excel_counts = Counter()
        excel_by_key = {}
        for d in despesas:
            desc_n = _norm(d["descricao"])
            pago_n = _norm(d["pago"])
            v_str  = _norm_valor(d["valor"])
            key = (desc_n, d["data"], v_str, pago_n)
            excel_counts[key] += 1
            excel_by_key[key] = d  # keep one representative record per key

        to_insert = []
        for key, excel_count in excel_counts.items():
            db_count = db_counts.get(key, 0)
            needed = excel_count - db_count

            if needed <= 0:
                stats["skipped"] += excel_count
                continue

            # Insert exactly 'needed' copies
            d = excel_by_key[key]
            for _ in range(needed):
                to_insert.append({
                    "descricao": key[0],
                    "obs": d.get("obs", ""),
                    "data": d["data"],
                    "pago": d["pago"],
                    "valor": d["valor"],
                    "source": "excel",
                })
            stats["skipped"] += (excel_count - needed)

        if to_insert:
            r = requests.post(
                f"{SUPABASE_URL}/rest/v1/despesas",
                headers={**HEADERS_SB, "Prefer": "return=minimal"},
                json=to_insert,
            )
            if r.status_code < 400:
                stats["inserted"] += len(to_insert)
                log.info(f"Inserted {len(to_insert)} new despesas")
            else:
                stats["errors"] += len(to_insert)
                log.error(f"  Insert error: {r.text[:300]}")
        else:
            log.info("Despesas: no new records to insert")

    # ── OUTROS ────────────────────────────────────────────────────────────────
    if outros:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/outros?select=id,cat,data,valor,source"
            f"&source=neq.erp_deleted",
            headers=HEADERS_SB,
        )
        existing_outros = resp.json() if resp.status_code == 200 else []

        db_counts_o = Counter()
        for e in existing_outros:
            key = (_norm(e["cat"]), e["data"], _norm_valor(e["valor"]))
            db_counts_o[key] += 1

        excel_counts_o = Counter()
        excel_by_key_o = {}
        for o in outros:
            key = (_norm(o["cat"]), o["data"], _norm_valor(o["valor"]))
            excel_counts_o[key] += 1
            excel_by_key_o[key] = o

        to_insert_outros = []
        for key, excel_count in excel_counts_o.items():
            db_count = db_counts_o.get(key, 0)
            needed = excel_count - db_count
            if needed <= 0:
                continue
            o = excel_by_key_o[key]
            for _ in range(needed):
                to_insert_outros.append({
                    "cat": o["cat"],
                    "data": o["data"],
                    "valor": o["valor"],
                    "source": "excel",
                })

        if to_insert_outros:
            r = requests.post(
                f"{SUPABASE_URL}/rest/v1/outros",
                headers={**HEADERS_SB, "Prefer": "return=minimal"},
                json=to_insert_outros,
            )
            if r.status_code < 400:
                log.info(f"Inserted {len(to_insert_outros)} new outros")
            else:
                log.error(f"  Insert error: {r.text[:300]}")
        else:
            log.info("Outros: no new records to insert")

    # ── SYNC REPORT ──────────────────────────────────────────────────────────
    log.info(f"=== SYNC REPORT: inserted={stats['inserted']}, "
             f"skipped={stats['skipped']}, errors={stats['errors']} ===")
    return stats

# === MAIN ===
def run_sync():
    try:
        log.info("=" * 50)
        log.info("=== Starting bidirectional sync ===")

        # 1. Auth
        token = get_graph_token()

        # 2. Download Excel
        content, site_id, item_id = download_excel(token)
        log.info(f"Downloaded {len(content)} bytes")

        # 3. Read ERP items pending sync to Excel
        erp_despesas, erp_outros, del_despesas, del_outros = read_erp_items()

        has_erp_changes = erp_despesas or erp_outros or del_despesas or del_outros

        # 4. If there are ERP changes, write them to Excel and upload
        if has_erp_changes:
            log.info("--- Writing ERP changes to Excel ---")
            wb = load_workbook(io.BytesIO(content))
            modified = write_to_excel(wb, erp_despesas, erp_outros, del_despesas, del_outros)

            if modified:
                output = io.BytesIO()
                wb.save(output)
                excel_bytes = output.getvalue()

                log.info("--- Uploading modified Excel to SharePoint ---")
                if upload_excel(token, site_id, item_id, excel_bytes):
                    content = excel_bytes
                    mark_erp_items_synced()
                else:
                    log.error("Upload failed, skipping mark as synced")
            else:
                log.info("No actual modifications needed in Excel")
                mark_erp_items_synced()
        else:
            log.info("No ERP changes pending")

        # 5. Parse Excel and sync to Supabase (Excel → Supabase)
        despesas, outros = parse_excel(content)
        if despesas or outros:
            sync_to_supabase(despesas, outros)

        log.info("=== Sync complete ===")
    except Exception as e:
        log.error(f"Sync failed: {e}", exc_info=True)

if __name__ == "__main__":
    log.info(f"Bidirectional sync worker started. Interval: {SYNC_INTERVAL} minutes")
    run_sync()

    import schedule
    schedule.every(SYNC_INTERVAL).minutes.do(run_sync)

    while True:
        schedule.run_pending()
        time.sleep(60)
